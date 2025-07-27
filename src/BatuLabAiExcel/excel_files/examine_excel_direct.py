#!/usr/bin/env python3
"""
Direct examination of Excel file using openpyxl
This shows what would be available through the MCP excel-mcp-server
"""

import openpyxl
from openpyxl.utils import range_boundaries
import json
from datetime import datetime

def examine_excel_file(file_path):
    """Examine Excel file and display its contents"""
    
    print("=" * 70)
    print(f"EXAMINING EXCEL FILE: {file_path}")
    print("=" * 70)
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 1. Get workbook metadata
        print("\n1. WORKBOOK METADATA")
        print("-" * 30)
        print(f"File path: {file_path}")
        print(f"Total worksheets: {len(workbook.worksheets)}")
        print(f"Active sheet: {workbook.active.title}")
        print(f"Worksheet names: {[ws.title for ws in workbook.worksheets]}")
        
        # 2. Examine each worksheet
        for sheet_index, worksheet in enumerate(workbook.worksheets, 1):
            print(f"\n{sheet_index}. WORKSHEET: '{worksheet.title}'")
            print("-" * 50)
            
            # Get used range
            if worksheet.max_row == 1 and worksheet.max_column == 1:
                if worksheet.cell(1, 1).value is None:
                    print("   Status: Empty worksheet")
                    continue
            
            used_range = f"A1:{openpyxl.utils.get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            print(f"   Used range: {used_range}")
            print(f"   Dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns")
            
            # Read all data from the worksheet
            data = []
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                         min_col=1, max_col=worksheet.max_column, 
                                         values_only=True):
                # Convert any datetime objects to strings for display
                formatted_row = []
                for cell in row:
                    if isinstance(cell, datetime):
                        formatted_row.append(cell.strftime('%Y-%m-%d %H:%M:%S'))
                    elif cell is None:
                        formatted_row.append('')
                    else:
                        formatted_row.append(str(cell))
                data.append(formatted_row)
            
            # Display the data
            print(f"\n   Data Content ({len(data)} rows):")
            for i, row in enumerate(data):
                if i == 0:
                    print(f"   HEADER: {' | '.join(row)}")
                else:
                    print(f"   ROW {i:2d}: {' | '.join(row)}")
                    
                # Limit display to first 15 rows to avoid overwhelming output
                if i >= 14:
                    remaining = len(data) - i - 1
                    if remaining > 0:
                        print(f"   ... and {remaining} more rows")
                    break
            
            # Show some specific cell examples
            print(f"\n   Sample Cell Values:")
            if worksheet.max_row >= 1 and worksheet.max_column >= 1:
                print(f"   A1: '{worksheet['A1'].value}'")
            if worksheet.max_row >= 2 and worksheet.max_column >= 1:
                print(f"   A2: '{worksheet['A2'].value}'")
            if worksheet.max_row >= 1 and worksheet.max_column >= 2:
                print(f"   B1: '{worksheet['B1'].value}'")
            if worksheet.max_row >= 2 and worksheet.max_column >= 2:
                print(f"   B2: '{worksheet['B2'].value}'")
        
        # 3. Summary information
        print(f"\n3. SUMMARY")
        print("-" * 20)
        total_cells = sum(ws.max_row * ws.max_column for ws in workbook.worksheets)
        non_empty_cells = 0
        
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and str(cell).strip():
                        non_empty_cells += 1
        
        print(f"Total cells in workbook: {total_cells}")
        print(f"Non-empty cells: {non_empty_cells}")
        print(f"Data density: {(non_empty_cells/total_cells*100):.1f}%")
        
        # 4. Demonstrate what MCP tools would return
        print(f"\n4. MCP TOOL SIMULATION")
        print("-" * 30)
        
        # Simulate get_workbook_metadata
        metadata = {
            "file_path": file_path,
            "worksheets": []
        }
        
        for ws in workbook.worksheets:
            ws_info = {
                "name": ws.title,
                "used_range": f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}",
                "max_row": ws.max_row,
                "max_column": ws.max_column
            }
            metadata["worksheets"].append(ws_info)
        
        print("get_workbook_metadata would return:")
        print(json.dumps(metadata, indent=2))
        
        # Simulate read_data_from_excel for first worksheet
        if workbook.worksheets:
            first_ws = workbook.worksheets[0]
            sample_data = []
            
            # Read first 5 rows and 5 columns as example
            max_rows = min(5, first_ws.max_row)
            max_cols = min(5, first_ws.max_column)
            
            for row in first_ws.iter_rows(min_row=1, max_row=max_rows, 
                                        min_col=1, max_col=max_cols, 
                                        values_only=True):
                formatted_row = []
                for cell in row:
                    if isinstance(cell, datetime):
                        formatted_row.append(cell.strftime('%Y-%m-%d'))
                    elif cell is None:
                        formatted_row.append(None)
                    else:
                        formatted_row.append(cell)
                sample_data.append(formatted_row)
            
            read_result = {
                "file_path": file_path,
                "worksheet_name": first_ws.title,
                "range": f"A1:{openpyxl.utils.get_column_letter(max_cols)}{max_rows}",
                "data": sample_data
            }
            
            print(f"\nread_data_from_excel('{first_ws.title}', 'A1:{openpyxl.utils.get_column_letter(max_cols)}{max_rows}') would return:")
            print(json.dumps(read_result, indent=2))
        
        workbook.close()
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found")
    except Exception as e:
        print(f"Error examining file: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 70)
    print("EXAMINATION COMPLETED")
    print("=" * 70)

if __name__ == "__main__":
    examine_excel_file("test.xlsx")